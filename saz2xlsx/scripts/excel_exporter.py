import os, shutil
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
from scripts.http_utils import sanitize_excel_str

def export_excel(rows: List[Dict[str, str]], out_path: str, base_url: Optional[str], include_time: bool, separate_by_url: bool, debug: bool=False) -> None:
    remark_col = '비고(매칭점수)' if debug else '비고'

    cols = ['경로', 'Method', '진단 URL']
    if debug:
        cols += ['메뉴명(추정)', '매칭점수']
    cols += ['파라미터', '진단결과']
    if include_time:
        cols.append('진단 시각')
    cols.append(remark_col)

    superset = ['경로','Method','진단 URL','메뉴명(추정)','매칭점수','파라미터','진단결과']
    if include_time:
        superset.append('진단 시각')
    superset.append('비고')

    df_all = pd.DataFrame(rows, columns=superset).map(sanitize_excel_str)
    if '비고' in df_all.columns and remark_col != '비고':
        df_all = df_all.rename(columns={'비고': remark_col})
    df = df_all.reindex(columns=cols)

    out_dir = os.path.dirname(os.path.abspath(out_path)) or '.'
    os.makedirs(out_dir, exist_ok=True)

    base_file_path = 'temp/temp.xlsx'
    if not os.path.exists('temp'):
        os.makedirs('temp')
    if not os.path.exists(base_file_path):
        wb = Workbook(); ws = wb.active; ws.title = '진단요약'; wb.save(base_file_path)
    shutil.copy(base_file_path, out_path)
    wb = load_workbook(out_path)

    if separate_by_url:
        grouped = {}
        for row in rows:
            domain = urlparse(row['진단 URL']).netloc
            grouped.setdefault(domain, []).append(row)

        if '진단요약' in wb.sheetnames:
            wb.remove(wb['진단요약'])

        for domain, data in grouped.items():
            ws = wb[domain] if domain in wb.sheetnames else wb.create_sheet(title=domain)
            ws.delete_rows(1, ws.max_row)
            ws.cell(row=1, column=1, value=f"URL : {domain}"); ws['A1'].font = Font(bold=True)

            df_domain_all = pd.DataFrame(data, columns=superset).map(sanitize_excel_str)
            if '비고' in df_domain_all.columns and remark_col != '비고':
                df_domain_all = df_domain_all.rename(columns={'비고': remark_col})
            df_domain = df_domain_all.reindex(columns=cols)

            for col_idx, header in enumerate(cols, 1):
                ws.cell(row=2, column=col_idx, value=header)
            for r_idx, row in enumerate(dataframe_to_rows(df_domain, index=False, header=False), start=3):
                ws.append(row)
    else:
        ws = wb.active; ws.title = '진단요약'; ws.delete_rows(1, ws.max_row)
        if base_url:
            ws.cell(row=1, column=1, value=f"URL : {base_url}"); ws['A1'].font = Font(bold=True)
            try: wb.title = base_url.replace('https://','').replace('http://','').replace('/','_')[:31]
            except Exception: pass

        for col_idx, header in enumerate(cols, 1):
            ws.cell(row=2, column=col_idx, value=header)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
            ws.append(row)

    for sheet in wb.worksheets:
        # 얇은 테두리
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # 굵은 테두리
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # 2번째 줄(헤더)에 굵은 테두리 적용
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thick_border

        # 3번째 줄 이후는 얇은 테두리 적용
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border

        # 열 너비 자동조정
        for col_cells in sheet.columns:
            maxlen = 8
            for cell in col_cells:
                if cell.value is None: 
                    continue
                l = len(str(cell.value))
                maxlen = max(maxlen, l)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(maxlen + 2, 80)


    wb.save(out_path); wb.close()
